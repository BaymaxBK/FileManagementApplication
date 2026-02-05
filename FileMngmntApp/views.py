from django.contrib.auth import authenticate, login
from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from django.db import connection,models
from FileMngmntApp.models import CustomTable,customFields,CustomTaskTable,CustomTaskFields,AssignedTaskRows
import re
import os
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse,HttpResponse
import json
from datetime import datetime
from math import ceil
import openpyxl
from io import BytesIO
from django.core.files.storage import FileSystemStorage
import tempfile
from openpyxl import load_workbook,Workbook
import pandas as pd
from django.conf import settings


def home(request):
    return render(request,'home.html')

def is_userAdmin(user):
    return user.is_authenticated and user.is_staff and not user.is_superuser
        

def is_normal_user(user):
    return user.is_authenticated and not user.is_staff and not user.is_superuser


def custom_login(request):
 
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        print("USER After Auth :",user)
        
        if user is not None:
            
            login(request, user)
            
            if user.is_superuser:
                return redirect('/admin/')

            elif user.is_staff:
                return redirect('admin_dashboard')
            
            else:
                return redirect('user_profile')
            
        else:
            messages.error(request, 'Invalid credentials')

    return render(request, 'login.html')


@login_required
@user_passes_test(is_userAdmin)
def user_admin(request):
    return render(request,'admin_dashboard.html')


@login_required
def upload_file(request):
    return render(request, 'upload.html')


@login_required
@user_passes_test(is_normal_user)
def user_profile(request):
    return render(request, 'user_profile.html')


# ADMIN VIEW'S

def sanitize_name(name):

    name = re.sub(r'\W+', '_', name.strip().lower())
    if not re.match(r'^[a-z_]', name):
        name = f"col_{name}"
    return name


@login_required
@user_passes_test(is_userAdmin)
def create_custom_Tasktable(request):

    if request.method == "POST":

        task_name = request.POST.get("taskName")
        userName = request.POST.get("assignedUser")
        TableId=request.POST.get("tableID")
        rowIDs=request.POST.get("rowIds","")

        print("TaskName > ",task_name) 
        print("UserName > ",userName) 
        print("TableId > ",TableId)
        print("RowId's > ",rowIDs)

        if not rowIDs:
            return JsonResponse({"success": False, "message": f"Please Select Atleast One Item !"})


        print("Row Id's ",rowIDs.split(","))
        print("TableId : ",TableId)

        custom_table = get_object_or_404(CustomTable, id=TableId)

        sql_fileds=list(custom_table.fields.values_list('field_name',flat=True))
        display_fileds=list(custom_table.fields.values_list('display_name',flat=True))
        sql_type=list(custom_table.fields.values_list('field_type',flat=True))

        sqlFldAndDisplyFld_lookup=dict(zip(sql_fileds,display_fileds))
        print("Fields Dict :> ",sqlFldAndDisplyFld_lookup)
        
        #BUID SQL QUERY
        fields_typesql=[]
        for field,sql_field,sqltype in zip(display_fileds,sql_fileds,sql_type):
            fields_typesql.append(f'"{sql_field}" {sqltype}')

        fields_sql_query=", ".join(fields_typesql)
        print("Fidels type : ",fields_sql_query)

        taskName_name=sanitize_name(task_name)

        create_sql = f"CREATE TABLE IF NOT EXISTS {taskName_name} (id INTEGER PRIMARY KEY, {fields_sql_query} );"
        print("SQL QUERY :",create_sql)
        with connection.cursor() as cursor:
            cursor.execute(create_sql)
        
        user=User.objects.get(username=userName)
        print("User Id ",user)
        taskTable=CustomTaskTable.objects.create(
             
             BaseTable=custom_table,
             TaskNameDisplay=task_name.strip(),
             TaskName=taskName_name,
             assigned_by = request.user,
             assigned_to=user

        )

        for field,sql_field,sqltype in zip(display_fileds,sql_fileds,sql_type):
            
            CustomTaskFields.objects.create(
                TaskTable = taskTable,
                display_name = field,
                field_name = sql_field,    # Safe name
                field_type = sqltype    
            )

        print("Table Name : ",custom_table.table_name)

        id_list = [int(x) for x in rowIDs.split(",") if x.strip().isdigit()]
        select_clause=["id"] + sql_fileds

        if id_list:
            placeholders = ",".join(["%s"] * len(id_list))
            sql = f'SELECT {", ".join(select_clause)} FROM "{custom_table.table_name}" WHERE id IN ({placeholders}) ORDER BY id'
            with connection.cursor() as cursor:
                cursor.execute(sql, id_list)
                rows = cursor.fetchall()  # or cursor.fetchall() to get all

            # UPDATE THE ASSINGED ROW TO 'ASSINGED ROW TABLE'
            with connection.cursor() as cursor:

                placeholders = ','.join(['%s'] * len(id_list))

                sql = f'UPDATE "{custom_table.table_name}" SET assigned_to = %s WHERE id IN ({placeholders})'
                params = [userName] + id_list
                cursor.execute(sql,params)

            print("Assingned row's are Updated ! ")

        else:
            rows = []

        if rows:

            # INSERT THE CUSTOM TASK TABLE ROW TO THE TASK TABLE
            columns=",".join(select_clause)
            placeholders = ", ".join(["%s"] * len(select_clause))
            insert_sql = f'INSERT INTO "{taskName_name}" ({columns}) VALUES ({placeholders})'
            with connection.cursor() as cursor:
                cursor.executemany(insert_sql,rows)

            # INSERT THE ASSINGED ROW TO 'AssignedTaskRows' Table 
            for row_id in id_list:
                AssignedTaskRows.objects.create(
                
                    task = taskTable,
                    assigned_from=custom_table,
                    assigned_row_id=row_id
                )
                

            return JsonResponse({"success": True, "message": f"Task '{task_name}' Assigned to {userName}"})
    
    return JsonResponse({"success": False, "message": "Invalid request method"})



@login_required
@user_passes_test(is_userAdmin)
def create_custom_table(request):

    if request.method == 'POST':
        
        table_name = request.POST.get('table_name')
        field_names = request.POST.getlist('field_name[]')
        field_types = request.POST.getlist('field_type[]')
        field_max_lengths = request.POST.getlist('max_length[]')

        not_nulls    = request.POST.getlist("not_null[]")
        uniques      = request.POST.getlist("unique[]")
        defaults     = request.POST.getlist("default_value[]")
        checks       = request.POST.getlist("check_condition[]")
        composite_pks = request.POST.getlist("composite_pk[]")

        print("Fields Name : ",field_names)
        print("Filed Types :",field_types)
        print("max size :",field_max_lengths)
        print("not null : ",not_nulls)
        print("uiqeue : ",uniques)
        print("default  : ",defaults)
        print("checks : ",checks)
        print("Composite pks : ",composite_pks)
        #return redirect('create_table')

        if not table_name or not field_names or not field_types or len(field_names) != len(field_types):
            messages.error(request, 'Invalid form data')
            return redirect('create_table')
        
        table_name_sql=sanitize_name(table_name)
        field_NameTypeConst_sql=[]
        
        Sql_types=["BOOLEAN","DATE","INTEGER","TEXT"]
        for i, name in enumerate(field_names):
            
            constraints = []
            safe_name=sanitize_name(name)

            type_= field_types[i] if field_types[i] in Sql_types  else "TEXT" 
            len_=field_max_lengths[i]

            sql_type=type_
            
            if type_ == "TEXT":
                try:
                    max_len = int(len_) if type_ else 255
                except ValueError:
                    max_len = 255

                sql_type = f"VARCHAR({max_len})"

            
            if str(i) in not_nulls:
                constraints.append("NOT NULL")
            if str(i) in uniques:
                constraints.append("UNIQUE")
            if defaults[i].strip():
                constraints.append(f"DEFAULT '{defaults[i].strip()}'")
            if checks[i].strip():
                constraints.append(f"CHECK ({checks[i].strip()})")

            field_NameTypeConst_sql.append(f'"{safe_name.strip()}" {sql_type} {" ".join(constraints)}')

            #column_defs.append(f'"{name.strip()}" {field_types[i]} {" ".join(constraints)}')

            # --- Composite Primary Key handling ---
            # pk_clause = ""
            # if use_composite_pk:
            #     pk_columns = []
            #     for i, name in enumerate(field_names):
            #         if str(i) in composite_pks:
            #             pk_columns.append(f'"{name.strip()}"')
            #     if not pk_columns:
            #         messages.error(request, "Select at least one column for the composite primary key.")
            #         return render(request, "create_table.html")
            #     pk_clause = f", PRIMARY KEY ({', '.join(pk_columns)})"
            # else:
            #     # if no composite PK, we still add a default 'id' column
            #     column_defs.insert(0, 'id SERIAL PRIMARY KEY')
    
        #     sql = f'CREATE TABLE "{table_name}" ({", ".join(column_defs)}{pk_clause})'
    
        #     try:
        #         with connection.cursor() as cursor:
        #             cursor.execute(sql)
        #         messages.success(request, f"Table {table_name} created successfully.")
        #         return redirect("table_list")
        #     except Exception as e:
        #         messages.error(request, f"Error creating table: {e}")
    

        # return render(request, "create_table.html")
        

        fields_sql_query=", ".join(field_NameTypeConst_sql)

        create_sql = f"CREATE TABLE IF NOT EXISTS {table_name_sql} (id SERIAL PRIMARY KEY, {fields_sql_query} , assigned_to VARCHAR(100) DEFAULT '');"
        print("SQL QUERY :",create_sql)

        #return redirect('create_table')

        try:
            with connection.cursor() as cursor:
                cursor.execute(create_sql)
            
            Tabel=CustomTable.objects.create(
                
                display_name =table_name.strip(),
                table_name = table_name_sql,
                created_by=request.user
            )

            Sql_types=["BOOLEAN","DATE","INTEGER","TEXT"]
            for i, name in enumerate(field_names):

                constraints = []
                
                type_= field_types[i] if field_types[i] in Sql_types  else "TEXT" 
                len_=field_max_lengths[i]

                sql_type=type_

                if type_ == "TEXT":
                    try:
                        max_len = int(len_) if type_ else 255
                    except ValueError:
                        max_len = 255

                    sql_type = f"VARCHAR({max_len})"


                _not_null =True if str(i) in not_nulls else False
                _default =defaults[i].strip() if defaults[i].strip() else None
                _unique=True if str(i) in uniques else False
                _check=checks[i].strip() if checks[i].strip() else None
                        
                customFields.objects.create(

                    table =Tabel, 
                    display_name =name.strip(),  # Original name
                    field_name = sanitize_name(name),# Safe name
                    field_type = sql_type,
                    max_length=int(max_len) if "VARCHAR" in sql_type and max_len else None,
                    is_primary_key=False,
                    is_not_null=_not_null,
                    is_unique=_unique,
                    default_value=_default,
                    check_constraint=_check

                )

            # for fname, ftype,max_len in zip(field_names, field_types,field_max_lengths):
                
            #     customFields.objects.create(

            #         table =Tabel, 
            #         display_name =fname.strip(),  # Original name
            #         field_name = sanitize_name(fname),# Safe name
            #         field_type = ftype,
            #         max_length=int(max_len) if "VARCHAR" in ftype and max_len else None
            #     )

            messages.success(request, f"Table '{table_name}' created successfully.")

        except Exception as e:
            
            messages.error(request, f"Error creating table: {str(e)}")

        return redirect('create_table')

    return render(request, 'create_table.html')



@login_required
def view_staff_created_models(request):

    staff_users = User.objects.filter(is_staff=True, is_superuser=False)
    tables = CustomTable.objects.prefetch_related('fields').filter(created_by__in=staff_users)

    return render(request, 'view_tables.html', {'tables': tables})


def view_table_data_updated(request,table_id):

    print("Task id  >> :",table_id)
    table = get_object_or_404(CustomTable, id=table_id)
    users = User.objects.filter(is_staff=False, is_superuser=False)
    is_admin=is_userAdmin(request.user)

    print("Table Name : >> ",table.display_name)
    return render(request, "view_table_data_updated.html", {"table": table,"users":users,"is_admin":is_admin})

def fetch_Table_data(request,table_id):
    
    """Return paginated JSON data for DataTables"""
    table = get_object_or_404(CustomTable, id=table_id)
    can_edit=table.can_user_edit(request.user)
    print(f'User "{request.user}" Can Edit > ',can_edit)
    
    
    table_name = table.table_name

    # SQL FILDES and Display SQL FIELDS Name's
    sql_fileds=["id"] + list(table.fields.values_list('field_name',flat=True))
    display_fileds=["ID"] + list(table.fields.values_list('display_name',flat=True))

    sqlFldAndDisplyFld_lookup=dict(zip(sql_fileds,display_fileds))
    print("Fields Dict :> ",sqlFldAndDisplyFld_lookup)

    draw = int(request.GET.get("draw", 1))
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))
    search_global = request.GET.get("search[value]", "").strip()
    selected_columns = request.GET.getlist('selected_columns[]', [])

    selectedRow=request.GET.getlist("selectedRow[]",[])
    print("Before Filter-Selected Row's >",selectedRow)
    
    selectedRow=[int(rowId) for rowId in selectedRow if rowId.isdigit()]
    print("After Filter-Selected Row's >",selectedRow)

    print("Selected Columns :",selected_columns)

    select_clause=', '.join(f'"{col}"' for col in selected_columns) if selected_columns else "*"
    
    # --- SELECT CLAUSE ---


    # Column Filter
    column_searches = {}
    

    i = 0
    while True:
        column_name = request.GET.get(f"columns[{i}][data]")
        if column_name is None:
            break

        search_val = request.GET.get(f"columns[{i}][search][value]")
        if search_val:
            column_searches[i] = search_val  # regex string from clien
        i += 1
    
    print("Column Filter Value :>> ",column_searches)

    
    # Build SQL with WHERE clauses
    where_clauses = []
    params = []


    # Global search
    if search_global:

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                [table_name]
            )
            all_cols = [r[0] for r in cursor.fetchall()]

        # build OR conditions on a safe subset (here we cast to text)
        or_parts = []
        for c in all_cols:
            or_parts.append(f'"{c}"::text ILIKE %s')
            params.append(f"%{search_global}%")
        if or_parts:
            where_clauses.append("(" + " OR ".join(or_parts) + ")")
        

     # Column-specific filters (these are regexes like ^(A|B)$ sent from client)
    # Convert them to safe LIKEs if possible. If you want to support regex on server,
    # you can use ~ operator in PostgreSQL, but prefer simpler approach:
    for col_idx, pattern in column_searches.items():
        # find the column name for this index
        col_name = request.GET.get(f"columns[{col_idx}][data]")
        if not col_name:
            continue
        # pattern may be ^(val1|val2)$ -> extract values
        # remove ^ and $ and split by |
        if pattern.startswith('^') and pattern.endswith('$'):
            inner = pattern[2:-2]
            print(f"Innter TExt {inner}")
            # unescape regex delimiters - this is simplistic; better to have client pass raw values instead
            values = inner.split('|')
            placeholders = ",".join(["%s"] * len(values))
            where_clauses.append(f'"{col_name}"::text IN ({placeholders})')
            params.extend(values)
        else:
            # fallback: ILIKE
            where_clauses.append(f'"{col_name}"::text ILIKE %s')
            params.append(f"%{pattern}%")

    if selectedRow:
        
        print("<< Selected Row ID >> ",selectedRow)
        placeholders = ",".join(["%s"] * len(selectedRow))
        where_clauses.append(f'id IN ({placeholders})')
        params.extend(selectedRow)

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""   
    

    # count total
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
        total = cursor.fetchone()[0]

    # count filtered
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT COUNT(*) FROM "{table_name}" {where_sql}', params)
        filtered = cursor.fetchone()[0]

    # fetch page
    length= length if length!=-1 else total
    final_query=f'SELECT {select_clause} FROM "{table_name}" {where_sql} ORDER BY id LIMIT %s OFFSET %s',params + [length, start]
    
    print(f"Param {params}")
    print("Final Query :> ",final_query)

    
    with connection.cursor() as cursor:
        cursor.execute(
            f'SELECT {select_clause} FROM "{table_name}" {where_sql} ORDER BY id LIMIT %s OFFSET %s',
            params + [length, start]
        )
        rows = cursor.fetchall()
        cols = [c[0] for c in cursor.description]

    data = [dict(zip(cols, r)) for r in rows]

    # ---------------- FILTER OPTIONS  ----------------
    filter_options = {}

    for col in selected_columns:

        with connection.cursor() as cursor:
            cursor.execute(
                f'''
                SELECT DISTINCT "{col}"
                FROM "{table_name}"
                {where_sql}
                ORDER BY "{col}"
                ''',
                params  # ❌ NO limit / offset
            )
            filter_options[col] = [
                row[0] if row[0] is not None else ''
                for row in cursor.fetchall()
            ]

    print("Filter Option :> ",filter_options)
    return JsonResponse({
        "draw": draw,
        "recordsTotal": total,
        "recordsFiltered": filtered,
        "data": data,
        "columns":cols,
        "displayColumn":sqlFldAndDisplyFld_lookup,
        "filter_options": filter_options,
        "can_edit":can_edit
    })

def update_table_cell(request):
    
    if request.method == "POST":
        
        taskTableId = request.POST.get("task_tableId")
        row_id = request.POST.get("id")
        field = request.POST.get("field")
        value = request.POST.get("value")

        custTblObj=get_object_or_404(CustomTable,id=taskTableId)
        taskName=custTblObj.table_name

        print(f"TASKTABLE Display {custTblObj.display_name} ROW ID> : {row_id}  FIELD> : {field}  VALUE> : {value}")

        if field == "id":
            print("< Trying to edit id column value ! >")
            return JsonResponse({"error":"Invalid DB Operation !"},status=400)
            
        # If empty, convert to NULL or empty string
        if value.strip() == "":
            value = None   # OR "" based on your requirement

        # UPDATE THE ASSINGED ROW TO 'ASSINGED ROW TABLE'
        with connection.cursor() as cursor:
            
            sql = f'UPDATE "{taskName}" SET "{field}" = %s WHERE id = %s'
            params = [value,row_id]
            cursor.execute(sql,params)


        return JsonResponse({"status": "success"},status=200)

    

@login_required
def user_assigned_tasks(request):
    # Fetch tasks assigned to the currently logged-in user
    assigned_tasks = CustomTaskTable.objects.filter(assigned_to=request.user)

    context = {
        'assigned_tasks': assigned_tasks
    }

    return render(request, 'user_assigned_tasks.html', context)


@login_required
def list_tables_for_user(request):
    user = request.user
    tables = CustomTable.objects.filter(
        models.Q(visible_to_users=user) |
        models.Q(visible_to_groups__in=user.groups.all())
    ).distinct()
    return render(request, "users_view_table.html", {"tables": tables})


@login_required
def list_assigned_tabletasks_for_admin(request,table_id):

    assigned_tasks = CustomTaskTable.objects.filter(BaseTable=table_id)

    context = {
        'assigned_tasks': assigned_tasks
    }
    
    return render(request, 'admin_view_assigned_task.html', context)


@login_required
def get_table_fields(request,table_id):
    
    print("GET TABLE FIELDS CALLED !")
    table=get_object_or_404(CustomTable,id=table_id)
    fields=list(table.fields.values_list('display_name',flat=True))

    print(f'fields : {fields}')
    return JsonResponse({"fields":fields})


@login_required
@user_passes_test(is_userAdmin)
def drop_tasktable(request,tasktable_id):

    task_table=get_object_or_404(CustomTaskTable,id=tasktable_id,assigned_by=request.user)
    baseTableId=task_table.BaseTable_id

    baseTableObj=get_object_or_404(CustomTable,id=baseTableId)

    print("BaseTable Name :> ",baseTableObj.table_name)
    delete_ids=[]

    try:

        with connection.cursor() as cursor:
            cursor.execute(f'SELECT id FROM "{task_table.TaskName}"')
            delete_ids = [row[0] for row in cursor.fetchall()]

        print("Assing Update Id's >> ",delete_ids)
        
    
        with connection.cursor() as cursor:
            cursor.execute(f'DROP TABLE IF EXISTS "{task_table.TaskName}"')            
        task_table.delete()  # Remove metadata
        
        
        with connection.cursor() as cursor:

                ids_placeholders = ','.join(['%s'] * len(delete_ids))
                sql = f'UPDATE "{baseTableObj.table_name}" SET "assigned_to" = \'\' WHERE id IN ({ids_placeholders})'
                if delete_ids:
                    cursor.execute(sql,delete_ids)

        messages.success(request, f"Task '{task_table.TaskNameDisplay}' Removed Successfully !.")
    except Exception as e:
        messages.error(request, f"Error deleting table: {str(e)}")

    return redirect('view_tables')

@login_required
def drop_table(request,table_id):

    table=get_object_or_404(CustomTable,id=table_id,created_by=request.user)
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"DROP TABLE IF EXISTS {table.table_name}")
        table.delete()  # Remove metadata

        messages.success(request, f"Table '{table.display_name}' deleted.")
    except Exception as e:
        messages.error(request, f"Error deleting table: {str(e)}")

    return redirect('view_tables')


def download_Tasktable_as_excel(request, task_id):
    
    # Fetch the custom table definition
    custom_Tasktable = get_object_or_404(CustomTaskTable, id=task_id)
    sql_fileds=list(custom_Tasktable.TaskFields.values_list('field_name',flat=True))
    display_fileds=list(custom_Tasktable.TaskFields.values_list('display_name',flat=True))

    sqlFldAndDisplyFld_lookup=dict(zip(sql_fileds,display_fileds))
    print("Fields Dict :> ",sqlFldAndDisplyFld_lookup)
    
    # Query table data dynamically
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT * FROM "{custom_Tasktable.TaskName}"')
        rows = cursor.fetchall()
        col_names_sql = [desc[0] for desc in cursor.description]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = custom_Tasktable.TaskName

    # Write headers
    col_names=[sqlFldAndDisplyFld_lookup.get(col_name) if col_name!="id" else "id" for col_name in col_names_sql ]

    print("Final Display Colname ",col_names)
    ws.append(col_names)

    # Write rows
    for row in rows:
        ws.append(row)

    # Build HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{custom_Tasktable.TaskNameDisplay}.xlsx"'

    wb.save(response)
    return response



def download_table_as_excel(request, table_id):
    
    # Fetch the custom table definition
    custom_table = get_object_or_404(CustomTable, id=table_id)
    sql_fileds=list(custom_table.fields.values_list('field_name',flat=True))
    display_fileds=list(custom_table.fields.values_list('display_name',flat=True))

    sqlFldAndDisplyFld_lookup=dict(zip(sql_fileds,display_fileds))
    print("Fields Dict :> ",sqlFldAndDisplyFld_lookup)
    
    # Query table data dynamically
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT * FROM "{custom_table.table_name}"')
        rows = cursor.fetchall()
        col_names_sql = [desc[0] for desc in cursor.description]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = custom_table.table_name

    # Write headers
    col_names=[sqlFldAndDisplyFld_lookup.get(col_name) if col_name!="id" else "id" for col_name in col_names_sql ]

    print("Final Display Colname ",col_names)
    ws.append(col_names)

    # Write rows
    for row in rows:
        ws.append(row)

    # Build HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{custom_table.table_name}.xlsx"'

    wb.save(response)
    return response


def update_excel_data(request, table_id):

    """Update existing table rows using Excel where the first column is 'id'."""
    custom_table = get_object_or_404(CustomTable, id=table_id)

    if request.method == "POST" and request.FILES.get("update_excel"):
        excel_file = request.FILES["update_excel"]

        # Save to a temp location
        fs = FileSystemStorage(location=settings.MEDIA_ROOT)
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        sql_fileds=list(custom_table.fields.values_list('field_name',flat=True))
        display_fileds=list(custom_table.fields.values_list('display_name',flat=True))

        DisplyFldAndsqlFld_lookup=dict(zip(display_fileds,sql_fileds))
        print("Fields Dict :> ",DisplyFldAndsqlFld_lookup)

        try:
            # UPDATE THE FILE DATA
            if request.POST.get('UpdateInsert') == "Update":
                
                print("File Update is Processing !")
                wb = load_workbook(file_path)
                ws = wb.active  # use first sheet

                headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                missing_header=[req_col for req_col in ["id"] + display_fileds if req_col not in headers]
                print("Missing Header :",missing_header)

                if missing_header:
                    messages.error(request, f"The Excel must contain following Columns : {', '.join(missing_header)}")
                    return redirect("view_tables")

                id_idx = [h.lower() for h in headers].index("id")

                # Collect rows as dicts
                update_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):

                    row_dict = {headers[i]: row[i] if row[i] is not None else '' for i in range(len(headers))}
                    update_rows.append(row_dict)

                with connection.cursor() as cursor:
                    cursor.execute(f'SELECT id FROM "{custom_table.table_name}"')
                    valid_ids = {row[0] for row in cursor.fetchall()}

                # Build SQL update  
                invalid_rows=[]
                with connection.cursor() as cursor:

                    for r in update_rows:          
                        row_id=r[headers[id_idx]]
                        if row_id not in valid_ids:
                            invalid_rows.append(r)
                            continue 

                        set_clause = ", ".join([f'"{DisplyFldAndsqlFld_lookup.get(df)}" = %s' for df in display_fileds])
                        sql = f'UPDATE "{custom_table.table_name}" SET {set_clause} WHERE id = %s'
                        values = [r[dsply_col] for dsply_col in display_fileds] + [row_id]
                        cursor.execute(sql, values)

                if invalid_rows:
                    print("Creating New Work BOOk!")
                    InvalidRow_wb = Workbook()
                    InvalidRow_ws = InvalidRow_wb.active
                    InvalidRow_ws.title = "Invalid IDs"

                    df = pd.read_excel(request.FILES["excel_file"])
                    # Add header row
                    print("Headers :",headers)
                    InvalidRow_ws.append(headers)

                    print("Header Were Created !")

                    # Add the invalid rows
                    for row in invalid_rows:
                        print("Row :",row)
                        row_data=list(row.values())
                        print("Row Data ",row_data)
                        InvalidRow_ws.append(row_data)

                    # 4️⃣ Create an HTTP response to download the file
                    response = HttpResponse(
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    response["Content-Disposition"] = f'attachment; filename="invalid_ids_{custom_table.table_name}.xlsx"'

                    InvalidRow_wb.save(response)

                    return response

                messages.success(request, "Table updated successfully.")
                
            #INSERT THE FILE DATA
            else:
                
                print("File Insertion is Processing !")
                df = pd.read_excel(file_path,keep_default_na=False)
                
                Missing_column=[]
                for col in display_fileds:
                    if col != "id"  and col not in df.columns:
                        Missing_column.append(col)

                if Missing_column:
                    messages.error(request, f"Missing Requried Column's: {', '.join(Missing_column)}")
                    return redirect('view_tables') 
                
                table=CustomTable.objects.get(id=table_id)

                fields=table.fields.all()
                print("Fields :>>",fields)
                
                db_table_name = table.table_name

                valid_rows, invalid_rows = [], []

                for idx, row in df.iterrows():
                    
                    row_dict = {}
                    error_msgs = []
        
                    for f in fields:
                        val = row.get(f.display_name)
        
                        # --- NOT NULL
                        if f.is_not_null and (pd.isna(val) or val == ""):
                            error_msgs.append(f"{f.display_name} is required")
                            continue
                        
                        # --- max_length
                        if f.max_length and isinstance(val, str) and len(val) > f.max_length:
                            error_msgs.append(f"{f.display_name} exceeds max length")
                            continue
                        
                        # --- default
                        if (val is None or val == "") and f.default_value:
                            val = f.default_value
        
                        row_dict[f.field_name] = val
        
                    if error_msgs:
                        invalid_rows.append({**row.to_dict(), "errors": "; ".join(error_msgs)})
                        continue
                    
                    # Check unique columns manually (to avoid DB errors first)
                    for f in fields.filter(is_unique=True):
                        with connection.cursor() as cur:
                            cur.execute(
                                f'SELECT 1 FROM "{db_table_name}" WHERE "{f.field_name}" = %s LIMIT 1',
                                [row_dict[f.field_name]],
                            )
                            if cur.fetchone():
                                invalid_rows.append({**row.to_dict(), "errors": f"{f.display_name} not unique"})
                                break
                    else:
                        valid_rows.append(row_dict)
        
                # --- Bulk insert valid rows
                if valid_rows:
                    columns = [f.field_name for f in fields]
                    values_sql = ",".join(
                        [
                            "(" + ",".join(["%s"] * len(columns)) + ")"
                            for _ in valid_rows
                        ]
                    )
                    params = [r[c] for r in valid_rows for c in columns]
                    insert_sql = f'INSERT INTO "{db_table_name}" ({",".join(columns)}) VALUES {values_sql}'
                    with connection.cursor() as cur:
                        cur.execute(insert_sql, params)
        
                # --- Return invalid rows as Excel if any
                if invalid_rows:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(list(invalid_rows[0].keys()))
                    for r in invalid_rows:
                        ws.append(list(r.values()))
        
                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    resp = HttpResponse(
                        buf,
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    resp["Content-Disposition"] = 'attachment; filename="invalid_rows.xlsx"'
                    return resp
                
                messages.success(request, f"Total Row's {len(valid_rows)} Records were Inserted successfully.")

        except Exception as e:
            messages.error(request, f"Error updating table: {e}")
        finally:
            os.remove(file_path)

    return redirect("view_tables")

def view_table_data(request, table_name):
    
    table = get_object_or_404(CustomTable, table_name=table_name)
    users = User.objects.filter(is_staff=False, is_superuser=False)


    row_limit=request.GET.get("limit",100)
    page_number=int(request.GET.get("page",1))
    display_columns=request.GET.get("displayColumn",[])
    

    with connection.cursor() as cursor:
        cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
        total_rows=cursor.fetchone()[0]

    print("Total No Row's :",total_rows)
    
    if row_limit != "all":
        row_limit=int(row_limit)
    else:
        row_limit=total_rows

    with connection.cursor() as cursor:
        cursor.execute(f'SELECT * FROM "{table_name}" ORDER BY id')
        columns = [col[0] for col in cursor.description]  # Get column names
        all_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]  # Convert each row to dict
    

    
    if display_columns:
        display_columns=display_columns.split(",")
    else:
        display_columns=columns

    print("Selected Column :",display_columns)

    #display_columns = [col for col in columns if col.lower() in selected_column]
    
    paginator=Paginator(all_rows,row_limit)
    page_obj=paginator.get_page(page_number)

    rows=list(page_obj.object_list)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":

        print("Row :",rows)
        print("Colum : ",display_columns)
        print("\n\n")
        print("No.of Pages :",paginator.num_pages)
        print("Curent Page : ",page_obj.number)

        return JsonResponse({

            "rows": rows,
            "columns":display_columns,
            'limit':row_limit,
            "total_pages": paginator.num_pages,
            "current_page": page_obj.number,
        
        })


    context = {
        'table': table,
        'columns': display_columns,
        'rows': rows,
        'limit':row_limit,
        'page_obj': page_obj,
        "total_pages": paginator.num_pages,
        "users":users
    }

    if is_normal_user(request.user):
        print("Is Normal User !")
        return render(request, 'userview_table_data.html', context)
    
    return render(request, 'view_table_data.html', context)


@csrf_exempt
def update_table_data(request, table_name):
    
    if request.method == 'POST':

        try:
            data = json.loads(request.body)
            rows = data.get('rows',[])
            print("Rows >>> ",rows)

            with connection.cursor() as cursor:
                  
                for row in rows:

                    row_id = int(row.pop('id',None))
                    print("ROW ",row_id)
                    # CHECK FOR DATE FORMATED DATA 
                    for column,data in row.items():
                        if isinstance(data,str) and "-" in data:
                            try:
                                parsed_date = datetime.strptime(data, "%d-%m-%Y").strftime("%Y-%m-%d")
                                row[column] = parsed_date
                            except ValueError:
                                pass  # Ignore if not a date                   
                    
                    if row_id:
                            
                        print(f"Row Id {row_id} >>> CONDTIONS {row_id > 0} ")
                        if row_id and int(row_id) > 0:    
   
                            #print("ROw id >> ", row_id)
                            #FORMATE THE DATE COLUMN DATA
                            for column,data in row.items():
                                if isinstance(data,str) and "-" in data:
                                    try:

                                        parsed_date = datetime.strptime(data, "%d-%m-%Y").strftime("%Y-%m-%d")
                                        row[column] = parsed_date

                                    except ValueError:
                                        pass  # Ignore if not a date


                            set_clauses = ', '.join([f'"{field}" = %s' for field in row.keys()])
                            values = list(row.values()) + [row_id]
                            #values.append(row_id)

                            sql = f'UPDATE "{table_name}" SET {set_clauses} WHERE id = %s'
                            cursor.execute(sql, values)

                        else:
                            
                            cursor.execute(f""" SELECT setval(
                                        pg_get_serial_sequence('"{table_name}"', 'id'),
                                        COALESCE(MAX(id), 0) + 1,
                                        false
                                        ) FROM "{table_name}"; """)                

                            column_clause = ", ".join([f'"{column}"' for column in row.keys()])  
                            column_data=list(row.values())
                            
                            value_clause = ", ".join(['%s']*len(row))
                            
                            Query_sql=f'INSERT INTO "{table_name}" ({column_clause}) VALUES ({value_clause})'

                            cursor.execute(Query_sql,column_data)
                
                # # ✅ Fix sequence after inserts (important for avoiding duplicate key errors)
                # cursor.execute(f"""
                #     SELECT setval(
                #         pg_get_serial_sequence('"{table_name}"', 'id'),
                #         COALESCE(MAX(id), 0) + 1,
                #         false
                #     ) FROM "{table_name}";
                # """)

            return JsonResponse({'success': True})
        
        except Exception as e:
            print("Update Error:", e)
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def view_row_data(request, table_name, row_id):

    table=get_object_or_404(CustomTable,table_name=table_name)
    print("Table In in View Row :", table.id)
    usercan_edit=table.can_user_edit(request.user)

    row={}
    print("VIEW CALLED ")
    with connection.cursor() as cursor:
        
        cursor.execute(f'SELECT * FROM {table_name} LIMIT 0')
        columns = [col[0] for col in cursor.description]

    with connection.cursor() as cursor:
        
        if int(row_id) > 0:
            cursor.execute(f'SELECT * FROM "{table_name}" WHERE id = %s', [row_id])
            columns = [col[0] for col in cursor.description]
            print("ROW ID :",row_id)
            row = dict(zip(columns, cursor.fetchone()))

    display_columns = [col for col in columns if col.lower() != 'id']

    print("Display Column :",display_columns)
    return render(request, "view_row_detail.html",{
        "row": row,
        "columns": columns,
        "table_id":table.id,
        "table_name":table_name,
        "is_userCanEdit":usercan_edit,
        "row_id":row_id
    })


@csrf_exempt
def delete_table_row(request, table_name):
    
    if request.method == "POST":

        print("Delete Row Hit .. !")
        try:
            data=json.loads(request.body)
            uniqueRowId=data.get("id")

            print(f"Data : {data}")
            with connection.cursor() as cursor:
                Delquery=f'DELETE FROM "{table_name}" WHERE id= %s'
                cursor.execute(Delquery, [uniqueRowId])
            return JsonResponse({'success': True})
        
        except Exception as e:
            return JsonResponse({"success":False,'error':str(e)})
    

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def adminViewdata_deleteSeleted_rows(request,table_name):
    ids = request.POST.getlist('ids[]')

    if not ids:
        return JsonResponse({'error': 'No IDs provided'}, status=400)

    with connection.cursor() as cursor:
        placeholders = ",".join(["%s"] * len(ids))
        cursor.execute(
            f'DELETE FROM "{table_name}" WHERE id IN ({placeholders})',
            ids
        )

    return JsonResponse({'success': True})

#USER VIEW'S
def view_taskTable_data(request, tasktable_id):

    """Render template with empty DataTable UI"""

    print("Task id :",tasktable_id)
    table = get_object_or_404(CustomTaskTable, id=tasktable_id)  
    return render(request, "assigned_task_data.html", {"table": table})


def fetch_taskTable_data(request, tasktable_id):

    """Return paginated JSON data for DataTables"""
    table = get_object_or_404(CustomTaskTable, id=tasktable_id)
    table_name = table.TaskName

    # SQL FILDES and Display SQL FIELDS Name's
    sql_fileds=["id"] + list(table.TaskFields.values_list('field_name',flat=True))
    display_fileds=["ID"] + list(table.TaskFields.values_list('display_name',flat=True))

    sqlFldAndDisplyFld_lookup=dict(zip(sql_fileds,display_fileds))
    print("Fields Dict :> ",sqlFldAndDisplyFld_lookup)

    draw = int(request.GET.get("draw", 1))
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))
    search_global = request.GET.get("search[value]", "").strip()
    selected_columns = request.GET.getlist('selected_columns[]', [])

    print("Selected Columns :",selected_columns)

    select_clause=', '.join(f'"{col}"' for col in selected_columns) if selected_columns else "*"
    
    # --- SELECT CLAUSE ---


    # Column Filter
    column_searches = {}
    

    i = 0
    while True:
        column_name = request.GET.get(f"columns[{i}][data]")
        if column_name is None:
            break

        search_val = request.GET.get(f"columns[{i}][search][value]")
        if search_val:
            column_searches[i] = search_val  # regex string from clien
        i += 1
    
    print("Column Filter Value :>> ",column_searches)

    
    # Build SQL with WHERE clauses
    where_clauses = []
    params = []


    # Global search
    if search_global:

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                [table_name]
            )
            all_cols = [r[0] for r in cursor.fetchall()]

        # build OR conditions on a safe subset (here we cast to text)
        or_parts = []
        for c in all_cols:
            or_parts.append(f'"{c}"::text ILIKE %s')
            params.append(f"%{search_global}%")
        if or_parts:
            where_clauses.append("(" + " OR ".join(or_parts) + ")")
        

     # Column-specific filters (these are regexes like ^(A|B)$ sent from client)
    # Convert them to safe LIKEs if possible. If you want to support regex on server,
    # you can use ~ operator in PostgreSQL, but prefer simpler approach:
    for col_idx, pattern in column_searches.items():
        # find the column name for this index
        col_name = request.GET.get(f"columns[{col_idx}][data]")
        if not col_name:
            continue
        # pattern may be ^(val1|val2)$ -> extract values
        # remove ^ and $ and split by |
        if pattern.startswith('^') and pattern.endswith('$'):
            inner = pattern[2:-2]
            print(f"Innter TExt {inner}")
            # unescape regex delimiters - this is simplistic; better to have client pass raw values instead
            values = inner.split('|')
            placeholders = ",".join(["%s"] * len(values))
            where_clauses.append(f'"{col_name}"::text IN ({placeholders})')
            params.extend(values)
        else:
            # fallback: ILIKE
            where_clauses.append(f'"{col_name}"::text ILIKE %s')
            params.append(f"%{pattern}%")

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""    

    # count total
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
        total = cursor.fetchone()[0]

    # count filtered
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT COUNT(*) FROM "{table_name}" {where_sql}', params)
        filtered = cursor.fetchone()[0]

    # fetch page
    length= length if length!=-1 else total
    final_query=f'SELECT {select_clause} FROM "{table_name}" {where_sql} ORDER BY id LIMIT %s OFFSET %s',params + [length, start]
    
    print(f"Param {params}")
    print("Final Query :> ",final_query)

    
    with connection.cursor() as cursor:
        cursor.execute(
            f'SELECT {select_clause} FROM "{table_name}" {where_sql} ORDER BY id LIMIT %s OFFSET %s',
            params + [length, start]
        )
        rows = cursor.fetchall()
        cols = [c[0] for c in cursor.description]

    data = [dict(zip(cols, r)) for r in rows]

    return JsonResponse({
        "draw": draw,
        "recordsTotal": total,
        "recordsFiltered": filtered,
        "data": data,
        "columns":cols,
        "displayColumn":sqlFldAndDisplyFld_lookup
    })

@csrf_exempt
def update_taskTable_cell(request):

    print("Update Taskcell Url Called !!")
    if request.method == "POST":
        taskTableId = request.POST.get("task_tableId")
        row_id = request.POST.get("id")
        field = request.POST.get("field")
        value = request.POST.get("value")

        custTskTblObj=get_object_or_404(CustomTaskTable,id=taskTableId)
        taskName=custTskTblObj.TaskName

        print(f"TASKTABLE Display {custTskTblObj.TaskNameDisplay} ROW ID> : {row_id}  FIELD> : {field}  VALUE> : {value}")

        # If empty, convert to NULL or empty string
        if value.strip() == "":
            value = None   # OR "" based on your requirement

        # UPDATE THE ASSINGED ROW TO 'ASSINGED ROW TABLE'
        with connection.cursor() as cursor:
            
            sql = f'UPDATE "{taskName}" SET "{field}" = %s WHERE id = %s'
            params = [value,row_id]
            cursor.execute(sql,params)


        return JsonResponse({"status": "success"})


@login_required
def user_task_update_status(request):

    taskId=request.POST.get("task_id")
    newStatus=request.POST.get("status")
    assigned_tasks = CustomTaskTable.objects.get(id=taskId)
    
    print("Assingned Tasks : ",assigned_tasks.TaskName)
    print("Tasks Status : ",newStatus)

    updated=CustomTaskTable.objects.filter(id=taskId).update(status=newStatus)
    if updated:
        return JsonResponse({"success":True,"message":f" Status Succesfully Updated to '{assigned_tasks.TaskNameDisplay}' Task"})
    
    return JsonResponse({'success': False, 'message': 'No update performed'})

@login_required
def user_assigned_tasks(request):
    # Fetch tasks assigned to the currently logged-in user
    assigned_tasks = CustomTaskTable.objects.filter(assigned_to=request.user)

    context = {
        'assigned_tasks': assigned_tasks
    }
    
    return render(request, 'user_assigned_tasks.html', context)



@login_required
def user_update_task_excel_data(request, tasktable_id):

    """Update existing table rows using Excel where the first column is 'id'."""
    custom_Tasktable = get_object_or_404(CustomTaskTable, id=tasktable_id)

    if request.method == "POST" and request.FILES.get("update_excel"):
        excel_file = request.FILES["update_excel"]

        # Save to a temp location
        fs = FileSystemStorage(location=settings.MEDIA_ROOT)
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        sql_fileds=list(custom_Tasktable.TaskFields.values_list('field_name',flat=True))
        display_fileds=list(custom_Tasktable.TaskFields.values_list('display_name',flat=True))

        DisplyFldAndsqlFld_lookup=dict(zip(display_fileds,sql_fileds))

        print("Fields Dict :> ",DisplyFldAndsqlFld_lookup)

        try:
            
            wb = load_workbook(file_path)
            ws = wb.active  # use first sheet

            headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            missing_header=[req_col for req_col in ["id"] + display_fileds if req_col not in headers]
            print("Missing Header :",missing_header)

            if missing_header:
                messages.error(request, f"The Excel must contain following Columns : {', '.join(missing_header)}")
                return redirect("view_tables")

            id_idx = [h.lower() for h in headers].index("id")

            # Collect rows as dicts
            update_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):

                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                update_rows.append(row_dict)

            with connection.cursor() as cursor:

                cursor.execute(f'SELECT id FROM "{custom_Tasktable.TaskName}"')
                valid_ids = {row[0] for row in cursor.fetchall()}

            # Build SQL update  
            invalid_rows=[]
            with connection.cursor() as cursor:

                for r in update_rows:          

                    row_id=r[headers[id_idx]]
                    if row_id not in valid_ids:
                        invalid_rows.append(r)
                        continue 

                    set_clause = ", ".join([f'"{DisplyFldAndsqlFld_lookup.get(df)}" = %s' for df in display_fileds])
                    sql = f'UPDATE "{custom_Tasktable.TaskName}" SET {set_clause} WHERE id = %s'
                    values = [r[dsply_col] for dsply_col in display_fileds] + [row_id]
                    cursor.execute(sql, values)

            if invalid_rows:
                
                print("Creating New Work BOOk!")
                InvalidRow_wb = Workbook()
                InvalidRow_ws = InvalidRow_wb.active
                InvalidRow_ws.title = "Invalid IDs"

                # df = pd.read_excel(request.FILES["excel_file"])
                # Add header row
                print("Headers :",headers)
                InvalidRow_ws.append(headers)
                print("Header Were Created !")
                
                # Add the invalid rows
                for row in invalid_rows:
                    print("Row :",row)
                    row_data=list(row.values())
                    print("Row Data ",row_data)
                    InvalidRow_ws.append(row_data)

                print("Invalid rows :> ",InvalidRow_ws)
                # 4️⃣ Create an HTTP response to download the file
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="invalid_ids_{custom_Tasktable.TaskName}.xlsx"'

                InvalidRow_wb.save(response)
                return response

            messages.success(request, "Table updated successfully.")
                
        except Exception as e:
            messages.error(request, f"Error updating table: {e}")
        finally:
            os.remove(file_path)

    return redirect("user_assigned_tasks")


# WORK FOR GLOBAL
def choose_table_and_upload(request):
    
    headers = []
    preview_rows = []
    rows_inserted = 0
    missing_headers = []
    extra_headers = []
    selected_table = None
    sheets=[]
    Selected_sheet=None

    # 
    
    user = request.user
    tables = CustomTable.objects.filter(
        models.Q(users_can_edit=user) &
        models.Q(visible_to_users=user)
    ).distinct()

    #tables = CustomTable.objects.all()

    try:
        if request.method == 'POST':
            table_id = request.POST.get('table_id')
            selected_table = get_object_or_404(CustomTable, id=table_id)

            print(f"Get FIle Condition : >> {request.FILES.get('excel_file')}")
            if request.FILES.get('excel_file'):

                excel_file = request.FILES['excel_file']
                fs = FileSystemStorage(location=settings.UPLOADS_DIR)
                filename = fs.save(excel_file.name, excel_file)

                file_path = fs.path(filename)

                # Save file path in session for later use
                # request.session['uploaded_excel_path'] = file_path
                request.session['uploaded_excel_path'] = f"uploads/{filename}"
                
            else:
                # No new file uploaded → try to get the saved one
                relative_path = request.session.get('uploaded_excel_path')
                print("Fetch From Session :(file Path ) ",relative_path)

                
                if not relative_path:
                    return render(request, 'choose_table.html', {
                        'tables': tables,
                        'error': 'No file uploaded or file missing.'
                    })
                
                file_path = os.path.join(settings.MEDIA_ROOT, relative_path)
                if not os.path.exists(file_path):
                    return render(request, 'choose_table.html', {
                            'tables': tables,
                            'error': 'File missing.'
                    })
                          
            print("Excel file :",file_path)

            # Load workbook
            wb = load_workbook(file_path,read_only=True)
            sheets=list(wb.sheetnames)

            Selected_sheet = request.POST.get('sheet_name')
            print("Selected Sheet >> ",Selected_sheet)
            if not Selected_sheet:
                Selected_sheet=sheets[0]

            Active_sheet =wb[Selected_sheet] 
            # Excel headers
            headers = [str(cell.value).strip() if cell.value else "" 
                       for cell in next(Active_sheet.iter_rows(min_row=1, max_row=1))]
            
            # Required headers from DB
            required_headers = list(selected_table.fields.values_list('display_name', flat=True))
            db_sanitized_headers=list(selected_table.fields.values_list('field_name', flat=True))

            # CREATE A LOOK UP
            header_lookup = dict(zip(required_headers, db_sanitized_headers))
            print("Header's And there Sanitized Name : ",header_lookup)

            # Check missing and extra headers
            missing_headers = [req for req in required_headers if req not in headers]
            
            extra_headers = [h for h in headers if h not in required_headers]
            
            # Only use matching headers
            matching_headers = [h for h in headers if h in required_headers]
            
            # Preview first 5 rows
            preview_rows = list(Active_sheet.iter_rows(min_row=2, values_only=True))[:5]
            print("Matching Headers : >>>> ",matching_headers)
            # Insert if confirmed and no missing headers

            if "confirm_insert" in request.POST and not missing_headers:
                
                print("Request POST >> ",request.POST)
                InsertOrupdateType=request.POST.get("InsertionOrUpdation_type")
                
                print("<< Type Instert Or Upated >> ",InsertOrupdateType)

                # --//----------------- INSERTION -----------------//--
                if InsertOrupdateType == "Insert":

                    print("File Insertion is Processing !")
                    df = pd.read_excel(file_path,sheet_name=Selected_sheet,keep_default_na=False)

                    Missing_column=[]
                    for col in required_headers:
                        if col != "id"  and col not in df.columns:
                            Missing_column.append(col)

                    if Missing_column:
                        messages.error(request, f"Missing Requried Column's: {', '.join(Missing_column)}")
                        return redirect('view_tables')
                    

                    fields=selected_table.fields.all()
            
                    db_table_name = selected_table.table_name

                    valid_rows, invalid_rows = [], []

                    # ITERATING THE DATAFRAME ROW'S  
                    for idx, row in df.iterrows():
                    
                        row_dict = {}
                        error_msgs = []

                        for f in fields:
                            val = row.get(f.display_name)

                            # --- NOT NULL
                            if f.is_not_null and (pd.isna(val) or val == ""):
                                error_msgs.append(f"{f.display_name} is required")
                                continue
                            
                            # --- max_length
                            if f.max_length and isinstance(val, str) and len(val) > f.max_length:
                                error_msgs.append(f"{f.display_name} exceeds max length")
                                continue
                            
                            # --- default
                            if (val is None or val == "") and f.default_value:
                                val = f.default_value

                            row_dict[f.field_name] = val

                        if error_msgs:
                            invalid_rows.append({**row.to_dict(), "errors": "; ".join(error_msgs)})
                            continue
                        
                        # Check unique columns manually (to avoid DB errors first)
                        for f in fields.filter(is_unique=True):
                            with connection.cursor() as cur:
                                cur.execute(
                                    f'SELECT 1 FROM "{db_table_name}" WHERE "{f.field_name}" = %s LIMIT 1',
                                    [row_dict[f.field_name]],
                                )
                                if cur.fetchone():
                                    invalid_rows.append({**row.to_dict(), "errors": f"{f.display_name} not unique"})
                                    break
                        else:
                            valid_rows.append(row_dict)
                    
                    print("Valid Data >> : ",valid_rows)
                    # INSERT A VALID ROWS
                    if valid_rows:
                        
                        columns = [f.field_name for f in fields]
                        values_sql = ",".join(
                            [
                                "(" + ",".join(["%s"] * len(columns)) + ")"
                                for _ in valid_rows
                            ]
                        )
                        params = [r[c] for r in valid_rows for c in columns]
                        insert_sql = f'INSERT INTO "{db_table_name}" ({",".join(columns)}) VALUES {values_sql}'
                        with connection.cursor() as cur:
                            cur.execute(insert_sql, params)
        
                    # --- Return invalid rows as Excel if any
                    if invalid_rows:
                        wb = Workbook()
                        ws = wb.active
                        ws.append(list(invalid_rows[0].keys()))
                        for r in invalid_rows:
                            ws.append(list(r.values()))

                        buf = BytesIO()
                        wb.save(buf)
                        buf.seek(0)
                        resp = HttpResponse(
                            buf,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                        resp["Content-Disposition"] = 'attachment; filename="invalid_rows.xlsx"'
                        return resp

                    messages.success(request, f"Total Row's {len(valid_rows)} Records were Inserted successfully.")
                    
                    # data_rows = []
                    # batch_size = 1000
                    
                    # sanitized_Matching_headers=[header_lookup[matched_header] for matched_header in matching_headers]
                    # print("Sanitized Headder in sql ",sanitized_Matching_headers)
    
                    # placeholders = ", ".join(["%s"] * len(sanitized_Matching_headers))
                    # col_names = ", ".join([f'"{col}"' for col in sanitized_Matching_headers])
                    # sql = f'INSERT INTO "{selected_table.table_name}" ({col_names}) VALUES ({placeholders})'
    
                    # for row in Active_sheet.iter_rows(min_row=2, values_only=True):
                    #     # Keep only matching columns
                    #     filtered_row = [row[headers.index(h)] for h in matching_headers]
                    #     data_rows.append(filtered_row)
                        
                    #     if data_rows:
                    #         print("Data Row >>>>>>>>>>> ",data_rows)
                    #         # Format date columns if needed
                    #         for r_index, r in enumerate(data_rows):
                    #             for c_index, val in enumerate(r):
                    #                 if isinstance(val, str) and "-" in val:
                    #                     try:
                    #                         data_rows[r_index][c_index] = datetime.strptime(val, "%d-%m-%Y").strftime("%Y-%m-%d")
                    #                     except ValueError:
                    #                         pass
                                        
                    #     if len(data_rows)>=batch_size:
                    #         with connection.cursor() as cursor:
                    #             cursor.executemany(sql, data_rows)
                    #         rows_inserted+=len(data_rows)
                    #         data_rows=[]
                    #         print("NO INSERTED ROW (BATCH ) :",rows_inserted)
    
                    # # LESS THAN BATCH LIMIT OF DATA 'OR' REMAINING DATA 
                    # if data_rows:
                    #     with connection.cursor() as cursor:
                    #         cursor.executemany(sql, data_rows)
                    #     rows_inserted += len(data_rows)
    
                    # print("Total Rows inserted:", rows_inserted)

                # --//----------------- UPDATION -----------------//--
                else:

                    print("<<<<<<  Updation To Be Done ! >>>>>>>")
                    print("File Update is Processing !")
                    
                    wb = load_workbook(file_path)
                    ws = wb.active  # use first sheet

                    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                    missing_header=[req_col for req_col in ["id"] + required_headers if req_col not in headers]
                    print("Missing Header :",missing_header)

                    if missing_header:
                        messages.error(request, f"The Excel must contain following Columns : {', '.join(missing_header)}")
                        return redirect("view_tables")

                    id_idx = [h.lower() for h in headers].index("id")

                    # Collect rows as dicts
                    update_rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):

                        row_dict = {headers[i]: row[i] if row[i] is not None else '' for i in range(len(headers))}
                        update_rows.append(row_dict)

                    with connection.cursor() as cursor:
                        cursor.execute(f'SELECT id FROM "{selected_table.table_name}"')
                        valid_ids = {row[0] for row in cursor.fetchall()}

                    # Build SQL update  
                    invalid_rows=[]
                    with connection.cursor() as cursor:

                        for r in update_rows:          
                            row_id=r[headers[id_idx]]
                            if row_id not in valid_ids:
                                invalid_rows.append(r)
                                continue 

                            set_clause = ", ".join([f'"{header_lookup.get(df)}" = %s' for df in required_headers])
                            sql = f'UPDATE "{selected_table.table_name}" SET {set_clause} WHERE id = %s'
                            values = [r[dsply_col] for dsply_col in required_headers] + [row_id]
                            cursor.execute(sql, values)

                    if invalid_rows:
                        print("Creating New Work BOOk!")
                        InvalidRow_wb = Workbook()
                        InvalidRow_ws = InvalidRow_wb.active
                        InvalidRow_ws.title = "Invalid IDs"
    
                        df = pd.read_excel(request.FILES["excel_file"])
                        # Add header row
                        print("Headers :",headers)
                        InvalidRow_ws.append(headers)
    
                        print("Header Were Created !")
    
                        # Add the invalid rows
                        for row in invalid_rows:
                            print("Row :",row)
                            row_data=list(row.values())
                            print("Row Data ",row_data)
                            InvalidRow_ws.append(row_data)
    
                        # 4️⃣ Create an HTTP response to download the file
                        response = HttpResponse(
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        response["Content-Disposition"] = f'attachment; filename="invalid_ids_{selected_table.table_name}.xlsx"'
    
                        InvalidRow_wb.save(response)
    
                        return response
    
                    messages.success(request, "Table updated successfully.")

    except Exception as e:
        print("Error:", (e))

    return render(request, 'choose_table.html', {
        'tables': tables,
        'headers': headers,
        'preview_rows': preview_rows,
        'selected_table': selected_table,
        'rows_inserted': rows_inserted,
        'missing_headers': missing_headers,
        'extra_headers': extra_headers,
        'sheets':sheets,
        'selected_sheet':Selected_sheet
    })


def get_sheet_names(request):

    sheet_names=[]
    
    print("Excel FIle Exists :",request.FILES.get("excel_file"))

    if request.method == 'POST' and request.FILES.get("excel_file"):
        
        print("SHEET NAME VIEW CALLED !")
        excel_file=request.FILES['excel_file']

        if not excel_file.name.endswith(('.xlsx', '.xlsm')):
            return JsonResponse({"error": "Only .xlsx or .xlsm files are supported"}, status=400)

        fs=FileSystemStorage(location=settings.UPLOADS_DIR)
        savefile_fs=fs.save(excel_file.name,excel_file)
        excel_file_path=fs.path(savefile_fs)

    
        # with tempfile.NamedTemporaryFile(delete=False) as temp:

        #     for chunk in excel_file.chunks():
        #         temp.write(chunk)
        #     temp_path=temp.name
        
        print("temp_path :>>",excel_file_path)
        

        workbook = load_workbook(excel_file_path, read_only=True)
        sheet_names = workbook.sheetnames

        Active_sheet=workbook.active
        columns = [cell for cell in next(Active_sheet.iter_rows(values_only=True))]

    
        print("SHEET NAME :",sheet_names)
        print("Active Sheet Column :",columns)
        workbook.close()

        return JsonResponse({"sheet_names":sheet_names,"activeSheet_columns":columns})
    
    return JsonResponse({"error": "No file uploaded"}, status=400)